from openpyxl import Workbook

from database import Appeal, get_all_appeal, session

async def create_excel_file(data: list[Appeal], path: str):
    wb = Workbook()
    sheet = wb.active

    sheet['A1'] = "id"
    sheet['B1'] = "tg_id"
    sheet['C1'] = "username"
    sheet['D1'] = "text"
    sheet['E1'] = "date"

    for i, appeal in enumerate(data):
        sheet.cell(row=i + 2, column=1).value = appeal.id
        sheet.cell(row=i + 2, column=2).value = appeal.tg_user_id
        sheet.cell(row=i + 2, column=3).value = appeal.tg_username
        sheet.cell(row=i + 2, column=4).value = appeal.appeal_text
        sheet.cell(row=i + 2, column=5).value = appeal.date

    wb.save(path)
